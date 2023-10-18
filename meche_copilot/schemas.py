"""
This module contains the schemas (used to validate the data) for the data objects used
"""

import io
import time
import ast
import tempfile
from langchain import PromptTemplate
from langchain.prompts.chat import (
    ChatPromptTemplate,
    SystemMessagePromptTemplate,
    AIMessagePromptTemplate,
    HumanMessagePromptTemplate,
)
import pandas as pd
from enum import Enum
from loguru import logger
from pathlib import Path
from datetime import datetime
from os.path import basename
from typing import ClassVar, List, Optional, Dict, Union, Any
from pydantic import BaseModel, root_validator, validator, Field, Extra
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.utils import get_column_letter

from meche_copilot.utils.config import load_config
from meche_copilot.utils.envars import PROJECT_ROOT

class EquipmentSpecificationAnalysis(BaseModel):
    """
    TODO (critical) - document this
    """
    equipment_name: str
    equipment_instance_design_uid: str
    spec_name: str
    spec_def: str
    spec_design_result: str
    spec_submittal_result: str
    spec_final_result: str
    notes: str = None

class SubmittalData(BaseModel):
    """
    TODO (critical) - document this
    """
    equipment_uid: str
    data: Dict[str, Any]

class EngineeringDesignSchedule(BaseModel):
    """
    The EngineeringDesignSchedule object contains the data relavent to a single schedule table in the engineering design document
    """
    equipment_name: Optional[str] = Field(description="Equipment name/type (eg. hydronic pump, exhaust fan, energy recovery ventilator")
    title: str = Field(description="Schedule table title")
    fpath: Path = Field(description="path to the pdf file")
    page_number: int = Field(description="page number of the schedule table")
    remarks: Optional[str] = Field(description="remarks or notes for the schedule table")
    headers: Optional[List[str]] = Field(description="schedule table column headers")
    row_data: Optional[Dict[str, List]] = Field(description="schedule table row data")

    class Config:
        extra = Extra.allow

class ScopeColumns(Enum):
    """
    The Scope document for a project that the user uploads (path specified in the session-config.yaml) is a spreadsheet with these required columns
    """
    num_instances = 0
    specifications_template_fpath = 1
    design_fpaths = 2
    design_notes = 3
    submittal_fpaths = 4
    submittal_notes = 5

class AgentConfig(BaseModel):
    """
    TODO (critical) - document this
    """
    system_prompt_template: SystemMessagePromptTemplate
    message_prompt_template: HumanMessagePromptTemplate
    model_name: Optional[str]

    # TODO - validate that the correct {{}} input keys for each prompt are present in the tempates provided in the config

    class Config:
        extra = 'allow'

    @validator('system_prompt_template', pre=True)
    def convert_to_system_prompt_template(cls, v):
        if isinstance(v, str):
            v = SystemMessagePromptTemplate.from_template(v)
            assert v is not None, f"Could not convert {v} to system message prompt template"
        return v
    
    @validator('message_prompt_template', pre=True)
    def convert_to_message_prompt_template(cls, v):
        if isinstance(v, str):
            v = HumanMessagePromptTemplate.from_template(v)
            assert v is not None, f"Could not convert {v} to human message prompt template"
        return v

class SpecResult(BaseModel):
    """
    The SpecResult object contains the data relavent to a single result for a spec (eg. the result for the spec for the pump's motor found on page, with value, has or han't been validated, and any annotations or comments about the result)
    """
    value: Optional[str]
    page: Optional[str]
    validation: Optional[str]
    annotation: Optional[str]

class SpecInstance(BaseModel):
    """
    The SpecInstance object contains the data relavent to a single instance of a spec (eg. the spec for the pump's motor)
    """
    name: str
    resA: SpecResult = SpecResult()
    resB: SpecResult = SpecResult()
    final_result: Optional[str] = None

class Source(BaseModel):
    """
    The Source object is the data source that is used to fill out the spec results for a piece of equipment (or example design documents are a Source and submittal documents are a Source)
    """
    name: str
    description: str
    ref_docs: Optional[List[Path]]
    notes: Optional[str]

    @validator('notes', 'ref_docs', pre=True)
    def check_length(cls, v):
        """If list is empty, set to None, if string is empty, set to None"""
        if len(v) == 0:
            v = None
        return v

    @validator('ref_docs', pre=True)
    def convert_to_list(cls, v):
        if isinstance(v, str):
            try:
                v = ast.literal_eval(v)
            except ValueError:
                v = [v]  # if the string cannot be evaluated, consider it as a single item list
        if isinstance(v, list):
            return [Path(item) for item in v]
        return v

class SessionConfig(BaseModel):
    """
    The SessionConfig object contains the data relavent to the session configuation which is loaded from the session-config yaml file
    """
    working_fpath: Path
    templates_fpath: Path
    scope_fpath: Path
    design: Source
    submittal: Source
    doc_retriever: AgentConfig
    spec_reader: AgentConfig
    spec_comparer: AgentConfig

    @validator('working_fpath', 'templates_fpath', 'scope_fpath')
    def file_must_exist(cls, v):
        logger.debug(f'Validating file path: {v}')
        fpath = PROJECT_ROOT / v
        assert fpath.exists(), f'The file {fpath} does not exist'
        return fpath
    
    @classmethod
    def from_yaml(cls, yaml_fpath: Path):
        logger.info(f'Loading config from {yaml_fpath}')
        config_box = load_config(yaml_fpath)
        return cls.parse_obj(config_box)

class ScopedEquipmentInstance(BaseModel):
    """
    The ScopedEquipmentInstance object contains the data relavent to a single instance of a piece of equipment (pump 1, pump 2, etc.)
    """
    name: str
    instances: List[SpecInstance]
    design_uid: Optional[str]
    design_data: Optional[Dict[str, str]]
    submittal_data: Optional[Dict[str, str]]

class ScopedEquipment(BaseModel):
    """
    A scoped equipment is a piece of equipment that has been scoped for the project

    Users add a scope excel file to whatever path is specified in the session config (eg. scope-fpath: ./data/demo-01/scope.xlsx) and the session object will be populated with the scoped equipment from that file.

    Eg.
        if the scoped equipment is a pump and the scope file says there are 6 instances of the pump then instances will be a list of 6 pump ScopedEquipmentInstance objects
    """
    name: str
    design_source: Source
    submittal_source: Source
    spec_defs: Dict[str, str] = {}
    instances: List[ScopedEquipmentInstance] = []

    # DEF_COL = "Definition"
    DEF_COL: ClassVar[str] = "Definition"

    class IOFormats(Enum):
        csv_str = "csv_str"
        df = "df"

    @validator("name")
    def validate_name(cls, value):
        """Validate name str isn't empty"""
        if not value:
            raise ValueError("name must not be empty")
        return value

    @validator("instances", "spec_defs", pre=True, each_item=False)
    def validate_instances_and_specs(cls, value, field):
        """Validate that length of instances and specs is > 1"""
        if field.name == "instances" and len(value) <= 1:
            raise ValueError("instances must contain more than 1 item")
        if field.name == "spec_defs" and len(value) <= 1:
            raise ValueError("spec_defs must contain more than 1 item")
        return value

    @validator("spec_defs", pre=True)
    def validate_spec_defs(cls, value):
        """Validate that spec_defs values are strings with lengths < MAX_SPEC_DEF_LEN"""
        MAX_SPEC_DEF_LEN = 500
        for key, val in value.items():
            if not len(val) < MAX_SPEC_DEF_LEN:
                raise ValueError(f"spec_defs value for key '{key}' must have a length less than {MAX_SPEC_DEF_LEN}. Got length of {len(val)} from: '{val}'")
        return value

    def spec_defs_to(self, output_format: IOFormats = IOFormats.csv_str) -> Union[pd.DataFrame, str]:
        """Convert spec_defs dict to specified output format"""

        assert self.spec_defs.keys(), "spec_defs must not be empty"

        df = pd.DataFrame.from_dict(self.spec_defs, orient='index', columns=[self.__class__.DEF_COL])

        # logger.debug(f"DF: {df}")
        # logger.debug(output_format.value)
        # logger.debug(self.IOFormats.df.value)
        # logger.debug(self.IOFormats.csv_str.value)
        
        if output_format.value == self.IOFormats.df.value:
            return df
        elif output_format.value == self.IOFormats.csv_str.value:
            csv_str = df.to_csv()
            return csv_str
        else:
            raise ValueError(f"This output format hasn't been implemented: {output_format}")
    
    @classmethod
    def spec_defs_from(cls, input: str, input_format: IOFormats = IOFormats.csv_str) -> Dict[str, str]:
        """Convert spec_defs str to specified output format"""

        if input_format.value == cls.IOFormats.df.value:
            return input.to_dict()[cls.DEF_COL]
        elif input_format.value == cls.IOFormats.csv_str.value:
            input_buffer = io.StringIO(input)
            df = pd.read_csv(filepath_or_buffer=input_buffer, index_col=0)
            return df.to_dict()[cls.DEF_COL]
        else:
            raise ValueError(f"This input format hasn't been implemented: {input_format}")
        

    def instances_to(self, output_format: IOFormats = IOFormats.csv_str) -> Union[pd.DataFrame, str]:
        data_rows = []
        row_labels = []
        for eq_inst in self.instances:
            row_labels.append(eq_inst.name)
            new_row = {}
            for spec_inst in eq_inst.instances:
                new_row[spec_inst.name] = "None"
            data_rows.append(new_row)
        df = pd.DataFrame(data_rows, index=row_labels)

        if output_format.value == self.IOFormats.df.value:
            return df
        elif output_format.value == self.IOFormats.csv_str.value:
            csv_str = df.to_csv()
            return csv_str
        else:
            raise ValueError(f"This output format hasn't been implemented: {output_format}")
    
    @classmethod
    def instances_from(cls, input: str, input_format: IOFormats = IOFormats.csv_str) -> List[ScopedEquipmentInstance]:

        if input_format.value == cls.IOFormats.csv_str.value:
            input_buffer = io.StringIO(input)
            df = pd.read_csv(
                filepath_or_buffer=input_buffer, 
                index_col=0
            )
        elif input_format.value == cls.IOFormats.df.value:
            df = input
        else:
            raise ValueError(f"This input format hasn't been implemented: {input_format}")

        eq_instances: List[ScopedEquipmentInstance] = []
        for index, row in df.iterrows():
            spec_instances = []
            for col_name, value in row.items():
                # Assuming the value is a tuple, you might need to parse it accordingly
                res = [None if x == 'None' else x for x in value.split(',')]
                if len(res) < 2:
                    res.append(None)
                spec_instances.append(SpecInstance(
                    name=col_name,
                    resA=SpecResult(value=res[0], page=res[1]),
                    resB=SpecResult(),
                ))
            eq_instances.append(ScopedEquipmentInstance(
                name=index,
                instances=spec_instances
            ))
        return eq_instances

class Session(BaseModel):
    """
    The Session object contains all the information needed to run the session/analysis
     
    Including: the session config params (from the yaml file), the datetime stamp of the start and last update of the session, and the list of scoped equipments
    """
    config: SessionConfig
    created_at: Optional[str] = None
    updated_at: Optional[str] = None
    equipments: List[ScopedEquipment] = []
    class Config:
        datetime_format = "%Y-%m-%d_%H.%M.%S"

    @classmethod
    def get_datetime_format(cls):
        return cls.Config.datetime_format
    
    @root_validator(pre=True)
    def set_timestamps(cls, values):
        datetime_format = cls.get_datetime_format()
        values["created_at"] = datetime.now().strftime(datetime_format)
        values["updated_at"] = datetime.now().strftime(datetime_format)
        return values

    @validator("*", pre=True)
    def update_timestamp(cls, value, values, field):
        datetime_format = cls.get_datetime_format()
        values["updated_at"] = datetime.now().strftime(datetime_format)
        return value

    @validator('equipments', pre=True, each_item=True)
    def validate_equipments(cls, v):
        """Number of equipment instances * number of specs shouldn't exceed MAX"""

        MAX_TABLE_SIZE = 100 # rows by columns (specs by instances)

        # TODO - this validation is a temp solution. I have no idea what this limit should be set to in terms of token processing....should prob just implement batching to deal with token limitations in the long term. Doing this for now.

        if len(v.spec_defs.keys()) * len(v.instances) > MAX_TABLE_SIZE:
            raise ValueError(f'The number of equipment instances * number of specs should not exceed {MAX_TABLE_SIZE}')
        return v

    @property
    def name(self) -> Optional[str]:
        if self.config.working_fpath:
            name = basename(self.config.working_fpath).replace('-', ' ').title()
            return name
        return None

    @classmethod
    def from_config(cls, config: SessionConfig):
        logger.info(f'Getting session info from config...')
        sess = cls(config=config)
        sess.load_equipments_from_scope()
        return sess

    def to_equipment_worksheet(self):
        logger.info(f'Writing equipment worksheet...')
        sheets_and_dfs = []

        for eq in self.equipments: # three sheets per equipment

            # equipment specs results worksheet
            spec_results_data = {
                "instance": [],
                "spec": [],
                "resA.value": [],
                "resA.page": [],
                "resA.validation": [],
                "resA.annotation": [],
                "resB.value": [],
                "resB.page": [],
                "resB.validation": [],
                "resB.annotation": [],
                "final_result": [],
            }
            for inst in eq.instances:
                # pd.json_normalize(inst["spec_results"]).T
                for spec in inst.instances:
                    spec_results_data["instance"].append(inst.name)
                    spec_results_data["spec"].append(spec.name)
                    spec_results_data["resA.value"].append(spec.resA.value)
                    spec_results_data["resA.page"].append(spec.resA.page)
                    spec_results_data["resA.validation"].append(spec.resA.validation)
                    spec_results_data["resA.annotation"].append(spec.resA.annotation)
                    spec_results_data["resB.value"].append(spec.resB.value)
                    spec_results_data["resB.page"].append(spec.resB.page)
                    spec_results_data["resB.validation"].append(spec.resB.validation)
                    spec_results_data["resB.annotation"].append(spec.resB.annotation)
                    spec_results_data["final_result"].append(spec.final_result)
            spec_results_df = pd.DataFrame(spec_results_data)

            # equipment-spec-defs worksheet
            spec_defs_df = pd.DataFrame.from_dict(eq.spec_defs, orient="index", columns=["description"])

            # equipment-info worksheet
            design_info = pd.DataFrame.from_dict(eq.design_source.dict(exclude_unset=True), orient="index", columns=["design_source"])
            submittal_info = pd.DataFrame.from_dict(eq.submittal_source.dict(exclude_unset=True), orient="index", columns=["submittal_source"])
            sources_df = pd.concat([design_info, submittal_info], axis=1)

            
            sheets_and_dfs.append((f"{eq.name}-sources", sources_df))
            sheets_and_dfs.append((f"{eq.name}-specs-defs", spec_defs_df))
            sheets_and_dfs.append((f"{eq.name}-specs-results", spec_results_df))
                
        output_fpath = self.config.working_fpath / "worksheet.xlsx"
        with pd.ExcelWriter(output_fpath) as writer:
            for sheet_name, df in sheets_and_dfs:
                index = False if sheet_name.endswith('results') else True
                df.to_excel(writer, sheet_name=sheet_name, index=index)
        logger.info(f'Wrote worksheet to {output_fpath}')
        return output_fpath

    def to_equipment_masterlist(self):
        logger.info(f'Writing equipment masterlist...')
        dfs = []
        for scoped_eq in self.equipments:
            sheet_name = scoped_eq.name
            instances = []
            for inst in scoped_eq.instances: 
                instance = {}
                for spec in inst.spec_results: # get spec.name and final result
                    print(f"{spec.name}: {spec.final_result}")
                    instance[spec.name] = spec.final_result
                instances.append(instance)
            scoped_eq_df = pd.DataFrame(instances)
            dfs.append((sheet_name, scoped_eq_df))

        output_fpath = self.config.working_fpath / "masterlist.xlsx"
        with pd.ExcelWriter(output_fpath) as writer:
            for sheet_name, df in dfs:
                df.to_excel(writer, sheet_name=sheet_name)
        return output_fpath

    def load_equipments_from_scope(self):
        logger.info(f'Getting scoped equipments...')
        scope_wb = load_workbook(str(PROJECT_ROOT / self.config.scope_fpath))
        scope_ws = scope_wb.active
        scoped_equipment: ScopedEquipment = []
        scope_columns = list(ScopeColumns.__members__.keys())
        for i, col_name in enumerate(scope_columns):
            if col_name != scope_ws[f"{get_column_letter(i+1)}1"].value:
                raise ValueError(f'The columns in the scope file do not match the fields in the ScopeColumns Enum')

        # each row in scope corresponds to a scoped equipment
        for row in scope_ws.iter_rows(min_row=2, values_only=True):
            name = Path(row[ScopeColumns.specifications_template_fpath.value]).stem.replace('-', ' ').replace('template', '').strip()
            template_fpath = self.config.templates_fpath / row[ScopeColumns.specifications_template_fpath.value]

            spec_defs = {}
            temp_wb = load_workbook(str(PROJECT_ROOT / template_fpath))
            temp_ws = temp_wb.active
            for col in temp_ws.iter_cols():
                spec_defs[col[0].value] = col[1].value or ""
            temp_wb.close()
            
            instances: List[ScopedEquipmentInstance] = []
            for i in range(1, row[ScopeColumns.num_instances.value]+1):
                instance_name = f"{name}-{i}"
                spec_results: List[SpecInstance] = []
                for spec_name in spec_defs.keys():
                    spec_results.append(SpecInstance(name=spec_name, resA=SpecResult(), resB=SpecResult(), final_result=None))
                instances.append(ScopedEquipmentInstance(name=instance_name, instances=spec_results))

            design_obj = Source(
                name=self.config.design.name,
                description=self.config.design.description,
                ref_docs=[ self.config.working_fpath / path.strip() for path in row[ScopeColumns.design_fpaths.value].split(',') ],
                notes=row[ScopeColumns.design_notes.value] or "",
                result=spec_results
            )
            submittal_obj = Source(
                name=self.config.submittal.name,
                description=self.config.submittal.description,
                ref_docs=[ self.config.working_fpath / path.strip() for path in row[ScopeColumns.submittal_fpaths.value].split(',') ],
                notes=row[ScopeColumns.submittal_notes.value] or "",
                result=spec_results
            )
            # instances = {'instances': instances} # wrap equipment instances in a dict to match the pydantic model
            scoped_equipment.append(ScopedEquipment(
                name=name,
                spec_defs=spec_defs,
                design_source=design_obj,
                submittal_source=submittal_obj,
                instances=instances
            ))
        self.equipments = scoped_equipment
        scope_wb.close()
    
    def update_from_worksheet(self):
        logger.info(f'Updating session from worksheet...')
        worksheet_fpath = self.config.working_fpath / 'worksheet.xlsx'
        new_equipments: List[ScopedEquipment] = []
        for eq in self.equipments:
            sources_df = pd.read_excel(worksheet_fpath, sheet_name=f'{eq.name}-sources', index_col=0)
            design = sources_df['design'].dropna().to_dict()
            submittal = sources_df['submittal'].dropna().to_dict()

            design_obj = Source(**design)
            submittal_obj = Source(**submittal)

            df_spec_defs = pd.read_excel(worksheet_fpath, sheet_name=f'{eq.name}-specs-defs', index_col=0)
            spec_defs = df_spec_defs.to_dict()['description']

            df_spec_results = pd.read_excel(worksheet_fpath, sheet_name=f'{eq.name}-specs-results')
            instances: ScopedEquipmentInstance = []
            inst_name = None
            for i, row in df_spec_results.iterrows():
                if row['instance'] != inst_name:
                    # append previous instance
                    if inst_name is not None:
                        instances.append(ScopedEquipmentInstance(**eq_instance))

                    # create new instance
                    inst_name = row['instance']
                    eq_instance = {}
                    eq_instance['name'] = inst_name
                    eq_instance['spec_results'] = []
                spec_res = {}
                spec_res['name'] = row['spec']
                spec_res['resA'] = {}
                spec_res['resA']['value'] = row['resA.value']
                spec_res['resA']['page'] = row['resA.page']
                spec_res['resA']['validation'] = row['resA.validation']
                spec_res['resA']['annotation'] = row['resA.annotation']
                spec_res['resB'] = {}
                spec_res['resB']['value'] = row['resB.value']
                spec_res['resB']['page'] = row['resB.page']
                spec_res['resB']['validation'] = row['resB.validation']
                spec_res['resB']['annotation'] = row['resB.annotation']
                spec_res['final_result'] = row['final_result']
                eq_instance['spec_results'].append(spec_res)

            scoped_eq = ScopedEquipment(
                name="pump",
                design_source=design_obj,
                submittal_source=submittal_obj,
                spec_defs=spec_defs,
                instances=instances,
            )
            
            new_equipments.append(scoped_eq)
        self.equipments = new_equipments
        
    def get_results(self):
        """Get results for each piece of equipment in the esd object"""
        
        specs_done = 0
        for eq in self.equipments:
            total_instances = len(eq.instances)
            total_specs = len(eq.spec_defs.keys())
            total_work = total_specs * total_instances
            try:
                # Calculate the time to sleep between each progress update
                sleep_time = 3.0 / total_work

                # TODO - run analyze specs chain for each instance
                # res = analyze_specs_chain.run({'equipments': [eq]}})
                # yield specs_done, f"Progress: {specs_done}/{total_work} specs done for {eq.name}"

                # Simulate the long-running function by sleeping and yielding progress updates
                for i in range(total_work):
                    # TODO - run analyze specs chain for each instance???
                    time.sleep(sleep_time)
                    specs_done += 1
                    yield specs_done, f"Progress: {specs_done}/{total_work} specs done for {eq.name}"
                
            except Exception as e:
                logger.exception(f"\nError getting results for {eq.name}\nMoving on...")

            specs_done += total_specs*total_instances
            yield specs_done, f"\nDone filling out worksheet for {eq.name}\n"